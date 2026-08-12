"""Phase 1 PMP import, reconciliation, metrics, and capacity domain services."""

import hashlib
import json
from collections import defaultdict
from datetime import date, datetime
from math import isclose
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import PmpArea, PmpImport, PmpImportError, PmpOrder, PmpOrderHistory, PmpPersonnel, PmpWeeklySchedule


JOSE_FILENAME = "JOSE.xlsx"
# JOSE.xlsx has five identical "Activo" headers.  These positions are the
# contract; never infer them from the repeated header labels.
JOSE_POSITIONS = {"external_id": 10, "area": 8, "planned_minutes": 14, "state": 15}
# FechaPlaneadaInicio is retained inside raw_payload_json because Phase 1 did
# not introduce an inferred creation-date column.  It is safe to use as an
# explicit planned-date cut, never as an order creation date.
JOSE_PLANNED_START_POSITION = 3
ALLOWED_STATES = {"ABIERTO": "pending", "FINALIZADO": "finalized"}
SHIFT_NAMES = {"1", "2", "3"}
PMP_TARGET_PERCENT = 90.0


def default_jose_path() -> Path:
    project_root = Path(__file__).resolve().parents[3]
    project_copy = project_root / "añadidos" / JOSE_FILENAME
    if project_copy.exists():
        return project_copy
    return Path(__file__).resolve().parents[1] / "data" / JOSE_FILENAME


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _payload(row: tuple[Any, ...]) -> str:
    return json.dumps({str(index): value for index, value in enumerate(row)}, default=str, ensure_ascii=False)


def _planned_start_from_source(value: Any) -> date | None:
    """Parse the explicit workbook date; never manufacture a creation date."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _totals(rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_area: dict[str, dict[str, Any]] = defaultdict(lambda: {"orders": 0, "planned_minutes": 0.0, "finalized_orders": 0, "pending_orders": 0, "finalized_minutes": 0.0, "pending_minutes": 0.0})
    for row in rows:
        target = by_area[row["area"]]
        target["orders"] += 1
        target["planned_minutes"] += row["planned_minutes"]
        target[f"{row['status']}_orders"] += 1
        target[f"{row['status']}_minutes"] += row["planned_minutes"]
    total = {"orders": len(rows), "planned_minutes": sum(row["planned_minutes"] for row in rows), "finalized_orders": sum(row["status"] == "finalized" for row in rows), "pending_orders": sum(row["status"] == "pending" for row in rows)}
    total["finalized_minutes"] = sum(row["planned_minutes"] for row in rows if row["status"] == "finalized")
    total["pending_minutes"] = sum(row["planned_minutes"] for row in rows if row["status"] == "pending")
    return {"global": total, "by_area": dict(sorted(by_area.items()))}


def _parse_rows(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    if path.name != JOSE_FILENAME:
        raise ValueError("La fuente inicial permitida es exclusivamente JOSE.xlsx")
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    valid: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        external_id = _text(row[JOSE_POSITIONS["external_id"]])
        area = _text(row[JOSE_POSITIONS["area"]]).upper()
        source_state = _text(row[JOSE_POSITIONS["state"]]).upper()
        raw_minutes = row[JOSE_POSITIONS["planned_minutes"]]
        error: tuple[str, str, str] | None = None
        if not external_id:
            error = ("external_id", "missing_external_id", "La orden no tiene identificador externo utilizable")
        elif not area:
            error = ("area", "missing_area", "Especialidad es obligatoria para el área PMP")
        elif source_state not in ALLOWED_STATES:
            error = ("state", "invalid_state", "Estado debe ser Abierto o Finalizado")
        else:
            try:
                planned_minutes = float(raw_minutes)
                if planned_minutes <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                error = ("planned_minutes", "invalid_planned_minutes", "TiempoPlaneado debe ser numérico y mayor que cero")
            else:
                if external_id in seen_ids:
                    error = ("external_id", "duplicate_external_id", "Identificador externo repetido en JOSE.xlsx")
        if error:
            invalid.append({"row_number": row_number, "field_name": error[0], "error_code": error[1], "error_message": error[2], "raw_payload_json": _payload(row)})
            continue
        seen_ids.add(external_id)
        valid.append({
            "row_number": row_number,
            "external_id": external_id,
            "area": area,
            "status": ALLOWED_STATES[source_state],
            "planned_minutes": planned_minutes,
            "planned_start_date": _planned_start_from_source(row[JOSE_PLANNED_START_POSITION]),
            "raw_payload_json": _payload(row),
        })
    return valid, invalid, worksheet.max_row - 1


def import_jose_workbook(db: Session, path: Path | None = None) -> PmpImport:
    source = path or default_jose_path()
    content_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    existing = db.query(PmpImport).filter(PmpImport.source_hash == content_hash).order_by(PmpImport.id.desc()).first()
    valid_rows, invalid_rows, total_rows = _parse_rows(source)
    imported = existing or PmpImport(source_filename=JOSE_FILENAME, source_hash=content_hash, status="completed")
    if not existing:
        db.add(imported)
        db.flush()
    imported.total_rows = total_rows
    imported.valid_rows = len(valid_rows)
    imported.invalid_rows = len(invalid_rows)
    imported.status = "completed"
    imported.reconciled_at = datetime.utcnow()
    areas: dict[str, PmpArea] = {}
    for name in sorted({row["area"] for row in valid_rows}):
        area = db.query(PmpArea).filter(PmpArea.name == name).first()
        if not area:
            area = PmpArea(name=name)
            db.add(area)
            db.flush()
        areas[name] = area
    existing_orders = {
        order.external_id: order
        for order in db.query(PmpOrder).filter(PmpOrder.pmp_import_id == imported.id).all()
    }
    for row in valid_rows:
        order = existing_orders.get(row["external_id"])
        if order:
            changed = any((
                order.pmp_area_id != areas[row["area"]].id,
                order.status != row["status"],
                order.planned_minutes != row["planned_minutes"],
                order.planned_start_date != row["planned_start_date"],
                order.source_row_number != row["row_number"],
            ))
            order.pmp_area_id = areas[row["area"]].id
            order.status = row["status"]
            order.planned_minutes = row["planned_minutes"]
            order.planned_start_date = row["planned_start_date"]
            order.source_row_number = row["row_number"]
            order.raw_payload_json = row["raw_payload_json"]
            if changed:
                db.add(PmpOrderHistory(pmp_order_id=order.id, action="excel_import_refresh", after_json=json.dumps({"import_id": imported.id, "row_number": row["row_number"]})))
            continue
        order = PmpOrder(external_id=row["external_id"], pmp_area_id=areas[row["area"]].id, status=row["status"], planned_minutes=row["planned_minutes"], planned_start_date=row["planned_start_date"], source="excel", source_row_number=row["row_number"], pmp_import_id=imported.id, raw_payload_json=row["raw_payload_json"])
        db.add(order)
        db.flush()
        db.add(PmpOrderHistory(pmp_order_id=order.id, action="excel_import", after_json=json.dumps({"import_id": imported.id, "row_number": row["row_number"]})))
    error_keys = set(db.query(PmpImportError.row_number, PmpImportError.field_name, PmpImportError.error_code).filter(PmpImportError.pmp_import_id == imported.id).all())
    for row in invalid_rows:
        key = (row["row_number"], row["field_name"], row["error_code"])
        if key not in error_keys:
            db.add(PmpImportError(pmp_import_id=imported.id, **row))
    _store_reconciliation(db, imported, valid_rows)
    db.commit()
    return imported


def import_summary(db: Session, imported: PmpImport) -> dict[str, Any]:
    errors = db.query(PmpImportError).filter(PmpImportError.pmp_import_id == imported.id).order_by(PmpImportError.row_number).all()
    reconciliation = reconcile_import(db, imported)
    return {"id": imported.id, "source_filename": imported.source_filename, "status": imported.status, "total_rows": imported.total_rows, "valid_rows": imported.valid_rows, "invalid_rows": imported.invalid_rows, "approved_at": imported.approved_at, "reconciliation": reconciliation, "errors": [{"row_number": item.row_number, "field_name": item.field_name, "code": item.error_code, "message": item.error_message} for item in errors]}


def reconcile_import(db: Session, imported: PmpImport) -> dict[str, Any]:
    """Reconcile DB rows against totals cached during the explicit import."""
    if not imported.reconciliation_json:
        return {
            "matches": False,
            "expected": {"global": {}, "by_area": {}},
            "persisted": _imported_totals(db, imported),
            "differences": {"global": {"reconciliation": "La carga debe ejecutarse una vez para almacenar su reconciliación."}, "by_area": {}},
        }
    try:
        expected = json.loads(imported.reconciliation_json)["expected"]
    except (KeyError, TypeError, json.JSONDecodeError):
        return {
            "matches": False,
            "expected": {"global": {}, "by_area": {}},
            "persisted": _imported_totals(db, imported),
            "differences": {"global": {"reconciliation": "La reconciliación persistida no es legible."}, "by_area": {}},
        }
    persisted = _imported_totals(db, imported)
    return _reconciliation_result(expected, persisted)


def _imported_totals(db: Session, imported: PmpImport) -> dict[str, Any]:
    actual_rows = db.query(PmpOrder, PmpArea).join(PmpArea).filter(PmpOrder.pmp_import_id == imported.id, PmpOrder.is_active.is_(True)).all()
    actual = [{"area": area.name, "status": order.status, "planned_minutes": order.planned_minutes} for order, area in actual_rows]
    return _totals(actual)


def _reconciliation_result(expected: dict[str, Any], persisted: dict[str, Any]) -> dict[str, Any]:
    differences = {"global": {}, "by_area": {}}
    for key in expected["global"]:
        delta = persisted["global"].get(key, 0) - expected["global"].get(key, 0)
        if delta:
            differences["global"][key] = delta
    for area in sorted(set(expected["by_area"]) | set(persisted["by_area"])):
        area_differences = {}
        for key in set(expected["by_area"].get(area, {})) | set(persisted["by_area"].get(area, {})):
            delta = persisted["by_area"].get(area, {}).get(key, 0) - expected["by_area"].get(area, {}).get(key, 0)
            if delta:
                area_differences[key] = delta
        if area_differences:
            differences["by_area"][area] = area_differences
    return {"matches": not differences["global"] and not differences["by_area"], "expected": expected, "persisted": persisted, "differences": differences}


def _store_reconciliation(db: Session, imported: PmpImport, expected_rows: list[dict[str, Any]]) -> None:
    expected = _totals(expected_rows)
    persisted = _imported_totals(db, imported)
    imported.reconciliation_json = json.dumps({"version": 1, "expected": expected, "persisted_at_import": persisted}, ensure_ascii=False)


def _status_color(compliant: bool, workload_percent: float) -> str:
    if compliant:
        return "green"
    if workload_percent >= 80:
        return "yellow"
    return "red"


def _format_metric(value: float | int) -> str:
    if isinstance(value, int) or isclose(float(value), round(float(value))):
        return f"{float(value):,.0f}"
    return f"{float(value):,.2f}"


def _metric_values(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total_orders = len(rows)
    finalized_orders = sum(row["status"] == "finalized" for row in rows)
    pending_orders = total_orders - finalized_orders
    planned_minutes = sum(float(row["planned_minutes"]) for row in rows)
    completed_minutes = sum(float(row["planned_minutes"]) for row in rows if row["status"] == "finalized")
    pending_minutes = planned_minutes - completed_minutes
    workload_completion_percent = round(100 * completed_minutes / planned_minutes, 2) if planned_minutes else 0.0
    order_completion_percent = round(100 * finalized_orders / total_orders, 2) if total_orders else 0.0
    compliant = workload_completion_percent > PMP_TARGET_PERCENT
    # This is intentionally a lower bound.  At exactly 90%, the bound is 0,
    # but a positive amount of work still has to be completed to satisfy >90%.
    strict_lower_bound = max(0.0, (PMP_TARGET_PERCENT / 100) * planned_minutes - completed_minutes)
    strict_note = (
        "Meta cumplida: el cumplimiento por carga es estrictamente mayor que 90 %."
        if compliant
        else "Debe completarse una cantidad estrictamente mayor a este límite para superar 90 %; 90 % exacto no cumple."
    )
    target_gap = round(max(0.0, PMP_TARGET_PERCENT - workload_completion_percent), 2)
    values: dict[str, Any] = {
        "total_orders": total_orders,
        "finalized_orders": finalized_orders,
        "pending_orders": pending_orders,
        "planned_minutes": round(planned_minutes, 2),
        "completed_planned_minutes": round(completed_minutes, 2),
        "pending_planned_minutes": round(pending_minutes, 2),
        "planned_hours": round(planned_minutes / 60, 2),
        "completed_planned_hours": round(completed_minutes / 60, 2),
        "pending_planned_hours": round(pending_minutes / 60, 2),
        "workload_completion_percent": workload_completion_percent,
        "order_completion_percent": order_completion_percent,
        "target_gap_percentage_points": target_gap,
        "additional_completed_minutes_lower_bound": round(strict_lower_bound, 2),
        "additional_completed_hours_lower_bound": round(strict_lower_bound / 60, 2),
        "strict_target_note": strict_note,
        "compliant": compliant,
        "traffic_light": _status_color(compliant, workload_completion_percent),
        "formatted": {},
    }
    values["formatted"] = {
        "total_orders": _format_metric(total_orders),
        "finalized_orders": _format_metric(finalized_orders),
        "pending_orders": _format_metric(pending_orders),
        "planned_hours": _format_metric(values["planned_hours"]),
        "completed_planned_hours": _format_metric(values["completed_planned_hours"]),
        "pending_planned_hours": _format_metric(values["pending_planned_hours"]),
        "workload_completion_percent": f"{workload_completion_percent:.2f}%",
        "order_completion_percent": f"{order_completion_percent:.2f}%",
    }
    return values


def _metric_values_from_aggregate(
    total_orders: int,
    finalized_orders: int,
    planned_minutes: float,
    completed_minutes: float,
) -> dict[str, Any]:
    """Apply the shared metric rules without materializing every order."""
    synthetic_rows = ([{"status": "finalized", "planned_minutes": completed_minutes}] if finalized_orders else [])
    pending_orders = total_orders - finalized_orders
    if pending_orders:
        synthetic_rows.append({"status": "pending", "planned_minutes": planned_minutes - completed_minutes})
    values = _metric_values(synthetic_rows)
    values["total_orders"] = total_orders
    values["finalized_orders"] = finalized_orders
    values["pending_orders"] = pending_orders
    values["order_completion_percent"] = round(100 * finalized_orders / total_orders, 2) if total_orders else 0.0
    values["formatted"]["total_orders"] = _format_metric(total_orders)
    values["formatted"]["finalized_orders"] = _format_metric(finalized_orders)
    values["formatted"]["pending_orders"] = _format_metric(pending_orders)
    values["formatted"]["order_completion_percent"] = f"{values['order_completion_percent']:.2f}%"
    return values


def metric_dictionary() -> dict[str, dict[str, str]]:
    """Single metric glossary returned with dashboard responses for UI tooltips."""
    return {
        "total_orders": {"label": "Órdenes totales", "formula": "count(órdenes PMP válidas en el alcance)", "source_fields": "pmp_orders.id, is_active"},
        "planned_hours": {"label": "Horas programadas", "formula": "sum(planned_minutes) / 60", "source_fields": "pmp_orders.planned_minutes (TiempoPlaneado)"},
        "completed_planned_hours": {"label": "Horas completadas", "formula": "sum(planned_minutes where status = finalized) / 60", "source_fields": "pmp_orders.status, planned_minutes"},
        "pending_planned_hours": {"label": "Horas pendientes", "formula": "sum(planned_minutes where status = pending) / 60", "source_fields": "pmp_orders.status, planned_minutes"},
        "workload_completion_percent": {"label": "Cumplimiento por carga", "formula": "completed_planned_minutes / planned_minutes × 100", "source_fields": "pmp_orders.status, planned_minutes"},
        "order_completion_percent": {"label": "Cumplimiento por órdenes", "formula": "finalized_orders / total_orders × 100", "source_fields": "pmp_orders.status"},
        "target": {"label": "Meta PMP", "formula": "workload_completion_percent > 90", "source_fields": "Regla de negocio PMP; 90 exacto no cumple"},
        "capacity": {"label": "Capacidad disponible", "formula": "sum(available_minutes de programación semanal activa) / 60", "source_fields": "pmp_weekly_schedules.available_minutes, pmp_personnel.is_active"},
    }


def _filtered_orders_query(
    db: Session,
    area_name: str | None = None,
    status: str | None = None,
    as_of_date: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
):
    """Build a SQL-only query over persisted operational PMP columns."""
    if status and status not in {"pending", "finalized"}:
        raise ValueError("El estado PMP debe ser pending o finalized")
    if as_of_date and (date_from or date_to):
        raise ValueError("Use fecha inicial/final o fecha de corte, no ambos")
    if date_from and date_to and date_from > date_to:
        raise ValueError("La fecha inicial no puede ser posterior a la fecha final")
    query = db.query(PmpOrder, PmpArea).join(PmpArea).filter(PmpOrder.is_active.is_(True))
    if area_name:
        query = query.filter(PmpArea.name == area_name.upper())
    if status:
        query = query.filter(PmpOrder.status == status)
    if as_of_date:
        query = query.filter(PmpOrder.planned_start_date <= as_of_date)
        mode = "planned_start_date_as_of"
    elif date_from or date_to:
        if date_from:
            query = query.filter(PmpOrder.planned_start_date >= date_from)
        if date_to:
            query = query.filter(PmpOrder.planned_start_date <= date_to)
        mode = "planned_start_date_range"
    else:
        mode = "not_requested"
    return query, {
        "order_date_filter": mode,
        "as_of_date": as_of_date.isoformat() if as_of_date else None,
        "date_from": date_from.isoformat() if date_from else None,
        "date_to": date_to.isoformat() if date_to else None,
        "orders_without_planned_start_date_excluded": 0,
    }


def dashboard_metrics(
    db: Session,
    area_name: str | None = None,
    status: str | None = None,
    as_of_date: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> dict[str, Any]:
    query, filter_application = _filtered_orders_query(db, area_name, status, as_of_date, date_from, date_to)
    area_query = db.query(PmpArea).filter(PmpArea.is_active.is_(True))
    if area_name:
        area_query = area_query.filter(PmpArea.name == area_name.upper())
    by_area: dict[str, dict[str, Any]] = {area.name: _metric_values([]) for area in area_query.order_by(PmpArea.name).all()}
    aggregates = (
        query.with_entities(
            PmpArea.name.label("area"),
            PmpOrder.status.label("status"),
            func.count(PmpOrder.id).label("orders"),
            func.coalesce(func.sum(PmpOrder.planned_minutes), 0.0).label("planned_minutes"),
        )
        .group_by(PmpArea.name, PmpOrder.status)
        .all()
    )
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in aggregates:
        grouped[row.area].append({"status": row.status, "planned_minutes": float(row.planned_minutes), "orders": int(row.orders)})
    for area, rows in grouped.items():
        total_orders = sum(row["orders"] for row in rows)
        finalized_orders = sum(row["orders"] for row in rows if row["status"] == "finalized")
        planned_minutes = sum(row["planned_minutes"] for row in rows)
        completed_minutes = sum(row["planned_minutes"] for row in rows if row["status"] == "finalized")
        by_area[area] = _metric_values_from_aggregate(total_orders, finalized_orders, planned_minutes, completed_minutes)
    ranking = sorted(by_area, key=lambda area: (-by_area[area]["pending_planned_minutes"], by_area[area]["workload_completion_percent"], area))
    for index, area in enumerate(ranking, start=1):
        by_area[area]["risk_rank"] = index
    total_orders = sum(row["orders"] for rows in grouped.values() for row in rows)
    finalized_orders = sum(row["orders"] for rows in grouped.values() for row in rows if row["status"] == "finalized")
    planned_minutes = sum(row["planned_minutes"] for rows in grouped.values() for row in rows)
    completed_minutes = sum(row["planned_minutes"] for rows in grouped.values() for row in rows if row["status"] == "finalized")
    global_metrics = _metric_values_from_aggregate(total_orders, finalized_orders, planned_minutes, completed_minutes)
    global_metrics["risk_rank"] = None
    high_risk_area = ranking[0] if ranking and by_area[ranking[0]]["pending_planned_minutes"] > 0 else None
    alerts = [] if global_metrics["compliant"] else [{
        "code": "pmp_compliance_below_target",
        "message": "El cumplimiento por carga PMP debe ser estrictamente mayor que 90 %; 90 % exacto no cumple.",
    }]
    if high_risk_area:
        alerts.append({"code": "highest_pending_workload", "message": f"{high_risk_area} concentra la mayor carga pendiente del alcance."})
    return {
        "global": global_metrics,
        "by_area": by_area,
        "area_ranking": ranking,
        "highest_risk_area": high_risk_area,
        "alerts": alerts,
        "filter_application": filter_application,
    }


def capacity_metrics(
    db: Session,
    week_start: date,
    metrics: dict[str, Any],
    area_name: str | None = None,
    shift_name: str | None = None,
) -> dict[str, Any]:
    """Compare area demand only when both real people and schedules are configured.

    Orders have no shift field, so a selected shift can show availability but
    cannot truthfully allocate pending demand or a staffing gap to that shift.
    """
    schedules = (
        db.query(PmpWeeklySchedule, PmpArea, PmpPersonnel)
        .join(PmpArea, PmpWeeklySchedule.pmp_area_id == PmpArea.id)
        .join(PmpPersonnel, PmpWeeklySchedule.pmp_personnel_id == PmpPersonnel.id)
        .filter(PmpWeeklySchedule.week_start == week_start, PmpPersonnel.is_active.is_(True))
    )
    if area_name:
        schedules = schedules.filter(PmpArea.name == area_name.upper())
    if shift_name:
        schedules = schedules.filter(PmpWeeklySchedule.shift_name == shift_name)
    schedule_rows = schedules.all()
    by_area: dict[str, list[tuple[PmpWeeklySchedule, PmpPersonnel]]] = defaultdict(list)
    shifts: set[str] = set()
    for schedule, area, person in schedule_rows:
        by_area[area.name].append((schedule, person))
        shifts.add(schedule.shift_name)
    result = []
    for area, area_metrics in metrics["by_area"].items():
        configured = by_area.get(area, [])
        if not configured:
            result.append({
                "area": area,
                "week_start": week_start.isoformat(),
                "shift": shift_name,
                "configured": False,
                "available_minutes": None,
                "available_hours": None,
                "required_minutes": None,
                "required_hours": None,
                "gap_minutes": None,
                "gap_hours": None,
                "staffing_gap_equivalent": None,
                "status": "not_configured",
                "missing_inputs": ["personal PMP activo", "programación semanal con minutos disponibles"],
            })
            continue
        available_minutes = sum(schedule.available_minutes for schedule, _ in configured)
        people_minutes: dict[int, float] = defaultdict(float)
        for schedule, person in configured:
            people_minutes[person.id] += schedule.available_minutes
        if shift_name:
            result.append({
                "area": area,
                "week_start": week_start.isoformat(),
                "shift": shift_name,
                "configured": True,
                "available_minutes": round(available_minutes, 2),
                "available_hours": round(available_minutes / 60, 2),
                "required_minutes": None,
                "required_hours": None,
                "gap_minutes": None,
                "gap_hours": None,
                "staffing_gap_equivalent": None,
                "status": "demand_not_assigned_to_shift",
                "missing_inputs": ["turno de cada orden PMP para asignar la demanda pendiente"],
            })
            continue
        required_minutes = area_metrics["pending_planned_minutes"]
        gap_minutes = max(0.0, required_minutes - available_minutes)
        average_minutes_per_scheduled_person = sum(people_minutes.values()) / len(people_minutes) if people_minutes else 0.0
        staffing_gap = gap_minutes / average_minutes_per_scheduled_person if average_minutes_per_scheduled_person else None
        result.append({
            "area": area,
            "week_start": week_start.isoformat(),
            "shift": None,
            "configured": True,
            "available_minutes": round(available_minutes, 2),
            "available_hours": round(available_minutes / 60, 2),
            "required_minutes": round(required_minutes, 2),
            "required_hours": round(required_minutes / 60, 2),
            "gap_minutes": round(gap_minutes, 2),
            "gap_hours": round(gap_minutes / 60, 2),
            "staffing_gap_equivalent": round(staffing_gap, 2) if staffing_gap is not None else None,
            "status": "insufficient" if gap_minutes > 0 else "sufficient",
            "missing_inputs": [],
        })
    return {
        "week_start": week_start.isoformat(),
        "configured": bool(schedule_rows),
        "rows": result,
        "available_shifts": sorted(shifts),
        "demand_shift_assignment_available": False,
        "notice": "La capacidad compara la carga pendiente por área. El turno de la orden no está disponible en la Fase 1, por lo que no se asigna demanda ni brecha por turno.",
    }


def list_orders(
    db: Session,
    area_name: str | None = None,
    status: str | None = None,
    as_of_date: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    offset: int = 0,
    limit: int = 50,
) -> dict[str, Any]:
    query, filter_application = _filtered_orders_query(db, area_name, status, as_of_date, date_from, date_to)
    total = query.order_by(None).count()
    page = query.order_by(PmpArea.name, PmpOrder.external_id, PmpOrder.id).offset(offset).limit(limit).all()
    items = [{
        "id": order.id,
        "external_id": order.external_id,
        "area": area.name,
        "status": order.status,
        "planned_minutes": float(order.planned_minutes),
        "planned_hours": round(order.planned_minutes / 60, 2),
        "planned_start_date": order.planned_start_date.isoformat() if order.planned_start_date else None,
        "source": order.source,
        "source_row_number": order.source_row_number,
    } for order, area in page]
    return {"items": items, "total": total, "offset": offset, "limit": limit, "filter_application": filter_application}


def validate_schedule(shift_name: str, available_minutes: float) -> None:
    if shift_name not in SHIFT_NAMES:
        raise ValueError("El turno PMP debe ser 1, 2 o 3")
    if available_minutes < 0:
        raise ValueError("Los minutos disponibles no pueden ser negativos")
